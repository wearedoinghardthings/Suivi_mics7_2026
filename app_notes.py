import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
import plotly.express as px
import io, qrcode, datetime, zipfile, base64, json, unicodedata

try:
    from supabase import create_client
    SUPABASE_LIB = True
except Exception:
    SUPABASE_LIB = False

try:
    from streamlit_qrcode_scanner import qrcode_scanner
    QR_OK = True
except Exception:
    QR_OK = False

try:
    from fpdf import FPDF
    PDF_OK = True
except Exception:
    PDF_OK = False

# ════════════════════════════════════════════════════════════
# PAGE CONFIG
# ════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="Formation — Plateforme",
    page_icon="📊", layout="wide",
    initial_sidebar_state="expanded"
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;500;600;700&family=DM+Serif+Display&display=swap');
html,body,[class*="css"]{font-family:'DM Sans',sans-serif!important;background:#fff!important;color:#1a1a2e!important;}
.main,.block-container{background:#fff!important;padding-top:.8rem!important;}
section[data-testid="stSidebar"]{background:linear-gradient(160deg,#f0f4ff,#e8f0fe)!important;border-right:1px solid #dce4f5;}
section[data-testid="stSidebar"] *{color:#1a1a2e!important;}
.main-title{font-family:'DM Serif Display',serif;font-size:clamp(1.4rem,4vw,2rem);color:#2c3e7a;}
.sub-title{font-size:clamp(.78rem,2.5vw,.9rem);color:#6b7db3;margin-bottom:1rem;}
.kpi-card{background:#fff;border:1px solid #e0e8ff;border-radius:14px;padding:clamp(10px,2vw,16px);
  text-align:center;box-shadow:0 2px 10px rgba(44,62,122,.06);margin-bottom:6px;}
.kpi-value{font-size:clamp(1.2rem,3.5vw,1.7rem);font-weight:700;color:#2c3e7a;line-height:1.1;}
.kpi-label{font-size:clamp(.6rem,1.8vw,.72rem);color:#7a8db3;margin-top:4px;
  text-transform:uppercase;letter-spacing:.6px;font-weight:500;}
.kpi-sub{font-size:.67rem;color:#a0aec0;margin-top:2px;}
.section-title{font-family:'DM Serif Display',serif;font-size:clamp(.95rem,3vw,1.2rem);
  color:#2c3e7a;margin-top:1.2rem;margin-bottom:.4rem;border-left:4px solid #4a6cf7;padding-left:11px;}
.stTabs [data-baseweb="tab-list"]{background:#f0f4ff;border-radius:10px;padding:4px;gap:3px;flex-wrap:wrap;}
.stTabs [data-baseweb="tab"]{border-radius:8px;font-weight:500;color:#2c3e7a;
  padding:5px 10px;font-size:clamp(.68rem,2vw,.82rem);white-space:nowrap;}
.stTabs [aria-selected="true"]{background:#fff!important;color:#4a6cf7!important;
  box-shadow:0 1px 6px rgba(74,108,247,.15);}
.stDownloadButton>button{background:linear-gradient(135deg,#4a6cf7,#7b5ea7)!important;
  color:#fff!important;border:none!important;border-radius:8px!important;font-weight:600!important;min-height:42px;}
.stButton>button{border-radius:8px!important;min-height:42px;}
.badge-tb{background:#e6f4ea;color:#1e7e34;padding:2px 9px;border-radius:20px;font-size:.76rem;font-weight:600;}
.badge-b{background:#dbeafe;color:#1a56db;padding:2px 9px;border-radius:20px;font-size:.76rem;font-weight:600;}
.badge-ab{background:#fef3c7;color:#92400e;padding:2px 9px;border-radius:20px;font-size:.76rem;font-weight:600;}
.badge-in{background:#fee2e2;color:#b91c1c;padding:2px 9px;border-radius:20px;font-size:.76rem;font-weight:600;}
.alerte-rouge{background:#fee2e2;border:1px solid #fca5a5;border-radius:10px;padding:10px 14px;margin:4px 0;}
.alerte-orange{background:#fef3c7;border:1px solid #fcd34d;border-radius:10px;padding:10px 14px;margin:4px 0;}
.profil-vert{background:#f0fdf4;border:1px solid #86efac;border-radius:10px;padding:8px 12px;font-size:.82rem;}
.profil-bleu{background:#eff6ff;border:1px solid #93c5fd;border-radius:10px;padding:8px 12px;font-size:.82rem;}
.profil-orange{background:#fff7ed;border:1px solid #fdba74;border-radius:10px;padding:8px 12px;font-size:.82rem;}
.profil-rouge{background:#fef2f2;border:1px solid #fca5a5;border-radius:10px;padding:8px 12px;font-size:.82rem;}
.db-ok{background:#f0fdf4;border:1px solid #bbf7d0;border-radius:10px;padding:10px 14px;margin-bottom:8px;font-size:.84rem;color:#166534;}
.db-warn{background:#fffbeb;border:1px solid #fcd34d;border-radius:10px;padding:10px 14px;margin-bottom:8px;font-size:.84rem;color:#92400e;}
.scan-ok{background:#f0fdf4;border:2px solid #86efac;border-radius:12px;padding:12px;text-align:center;margin:8px 0;}
.locked-box{background:#f8faff;border:2px dashed #c7d2fe;border-radius:14px;padding:40px;text-align:center;margin:20px 0;}
.qr-grid{display:flex;flex-wrap:wrap;gap:10px;}
.qr-item{background:#fff;border:1px solid #e0e8ff;border-radius:10px;padding:8px;
  text-align:center;width:clamp(110px,22vw,140px);}
.qr-name{font-size:.7rem;color:#2c3e7a;font-weight:600;margin-top:5px;line-height:1.3;word-break:break-word;}
@media(max-width:768px){
  .block-container{padding:.4rem .5rem 2rem!important;}
  .qr-item{width:calc(50% - 10px)!important;}
}
</style>
""", unsafe_allow_html=True)

# ════════════════════════════════════════════════════════════
# DONNÉES PAR DÉFAUT
# ════════════════════════════════════════════════════════════
DEFAULT_NOTES = {
    "Nom et Prénom":[
        "Acka Gustave Boris","Agnero Djadja","Kra Yao Assouman Aubin Gael",
        "Kouao Marie-Paule","Sangaré Aissata","Kouassi Kangah Akissi Philomène",
        "Yao Julienne","Amankan Adjo Josiane","Fouati Edwige","Agnehoura Vanessa",
        "Djeha Somaud Marthe Abigail","Gnomblé Christelle","Loua Saty Veronique",
        "Coulibaly Lancina Medard","Deza Boga Stefan","Beugré Brigitte",
        "Bogui Kacohon Prisca Sandrine epse Gris","Touré Madongon Mariam",
        "Konan Amenan Grâce-Victoire","Diarrassouba Madoussou",
    ],
    "Note 2":[18,18,17,18,17,13,16,16,15,14,14,15,0,14,0,10,0,14,0,0],
    "Note 3":[14,14,6,0,11,11,11,9,9,9,11,6,14,6,6,9,3,11,0,0],
    "Note 4":[15,18,15,14,13,14,13,15,14,16,0,14,15,9,10,0,15,15,11,0],
    "Note 5":[20,13,10,17,13,10,10,0,13,0,13,13,13,10,13,13,10,10,10,0],
    "Note 6":[12,16,16,12,14,16,16,16,12,14,14,8,12,12,12,18,14,12,12,10],
    "Note 7":[19,17,19,19,18,18,17,15,18,18,16,19,17,17,17,15,18,15,14,11],
    "Note 8":[20,20,18,18,15,18,13,17,13,17,17,15,15,15,17,12,18,0,13,10],
    "Note 9":[18,18,19,20,15,14,13,19,12,17,17,9,12,13,16,13,0,0,8,3],
}
DEFAULT_SESSIONS = ["Jour 1 - Matin","Jour 1 - Après-midi","Jour 2 - Matin",
                    "Jour 2 - Après-midi","Jour 3 - Matin"]
MOTIFS = ["Maladie","Raison familiale","Mission professionnelle","Autre","Sans justification"]

# ════════════════════════════════════════════════════════════
# SUPABASE
# ════════════════════════════════════════════════════════════
@st.cache_resource
def get_sb():
    if not SUPABASE_LIB: return None
    try:
        url = st.secrets["supabase"]["url"]
        key = st.secrets["supabase"]["key"]
        if "VOTRE" in url or not url.startswith("http"): return None
        return create_client(url, key)
    except: return None

sb = get_sb()
DB = sb is not None

@st.cache_data(ttl=15)
def db_get(cle):
    try:
        r = sb.table("app_config").select("valeur").eq("cle", cle).execute()
        if r.data: return r.data[0]["valeur"]
    except: pass
    return None

def db_set(cle, valeur):
    try:
        sb.table("app_config").upsert({"cle":cle,"valeur":valeur,
            "updated_at":datetime.datetime.utcnow().isoformat()}).execute()
        db_get.clear()
        return True
    except: return False

@st.cache_data(ttl=10)
def db_get_seuil(cle):
    try:
        r = sb.table("config_seuils").select("valeur").eq("cle", cle).execute()
        if r.data: return r.data[0]["valeur"]
    except: pass
    return None

def db_set_seuil(cle, valeur):
    try:
        sb.table("config_seuils").upsert({"cle":cle,"valeur":valeur}).execute()
        db_get_seuil.clear()
        return True
    except: return False

@st.cache_data(ttl=8)
def db_get_presences():
    try:
        r = sb.table("presences").select("*").order("date_jour").execute()
        if r.data:
            df = pd.DataFrame(r.data)
            df["date_jour"] = pd.to_datetime(df["date_jour"]).dt.date
            return df
    except: pass
    return pd.DataFrame(columns=["id","agent","date_jour","heure","session","statut"])

@st.cache_data(ttl=15)
def db_get_calendrier():
    try:
        r = sb.table("calendrier").select("*").order("date_jour").execute()
        if r.data:
            df = pd.DataFrame(r.data)
            df["date_jour"] = pd.to_datetime(df["date_jour"]).dt.date
            return df
    except: pass
    return pd.DataFrame(columns=["id","date_jour","a_une_note","nom_note","ordre_note"])

@st.cache_data(ttl=10)
def db_get_justifications():
    try:
        r = sb.table("justifications").select("*").execute()
        if r.data:
            df = pd.DataFrame(r.data)
            df["date_jour"] = pd.to_datetime(df["date_jour"]).dt.date
            return df
    except: pass
    return pd.DataFrame(columns=["id","agent","date_jour","session","type_absence","motif"])

def db_add_presence(agent, date_jour, heure, session, statut="Présent"):
    try:
        sb.table("presences").upsert({
            "agent":agent,"date_jour":str(date_jour),
            "heure":str(heure),"session":session,"statut":statut,
        }, on_conflict="agent,date_jour,session").execute()
        db_get_presences.clear()
        return True
    except: return False

def db_check_pointed(agent, session, date_jour=None):
    if date_jour is None: date_jour = datetime.date.today()
    try:
        r = sb.table("presences").select("heure")\
            .eq("agent",agent).eq("session",session)\
            .eq("date_jour",str(date_jour)).execute()
        if r.data: return True, r.data[0]["heure"]
    except: pass
    return False, None

def db_add_justification(agent, date_jour, session, type_abs, motif):
    try:
        sb.table("justifications").upsert({
            "agent":agent,"date_jour":str(date_jour),
            "session":session,"type_absence":type_abs,"motif":motif,
        }, on_conflict="agent,date_jour,session").execute()
        db_get_justifications.clear()
        return True
    except: return False

def db_save_calendrier(rows):
    try:
        sb.table("calendrier").delete().neq("id",0).execute()
        for row in rows:
            sb.table("calendrier").insert(row).execute()
        db_get_calendrier.clear()
        return True
    except: return False

# ════════════════════════════════════════════════════════════
# MÉMOIRE FALLBACK
# ════════════════════════════════════════════════════════════
@st.cache_resource
def mem():
    return {"notes":None,"sessions":DEFAULT_SESSIONS,"presences":[],
            "justifications":[],"calendrier":[],"qr_codes":{},
            "notes_visibles":False,"notes_name_col":None}
M = mem()

def _detect_name_col(df_raw):
    """
    Détecte la colonne des noms = celle avec le PLUS de valeurs non numériques.
    Fonctionne même si les colonnes sont dans le désordre.
    """
    best_col = df_raw.columns[0]
    best_score = -1
    for c in df_raw.columns:
        vals = df_raw[c].astype(str).str.replace(",",".").str.strip()
        non_num = vals.apply(lambda x: pd.to_numeric(x, errors='coerce')).isna().sum()
        if non_num > best_score:
            best_score = non_num
            best_col = c
    return best_col

def get_notes():
    if DB:
        v = db_get("notes")
        df_raw = pd.DataFrame(v) if v else pd.DataFrame(DEFAULT_NOTES)
    else:
        df_raw = pd.DataFrame(M["notes"]) if M["notes"] else pd.DataFrame(DEFAULT_NOTES)

    # Détecter la colonne des noms AVANT toute transformation
    name_c = _detect_name_col(df_raw)

    # Convertir chaque colonne correctement
    for c in df_raw.columns:
        if c == name_c:
            # Colonne des noms → garder comme texte
            df_raw[c] = df_raw[c].astype(str).str.strip()
        else:
            # Colonne de notes → remplacer virgules et convertir en nombre
            df_raw[c] = pd.to_numeric(
                df_raw[c].astype(str).str.replace(",",".").str.strip(),
                errors="coerce"
            ).fillna(0)

    # Remettre la colonne des noms EN PREMIER
    cols = [name_c] + [c for c in df_raw.columns if c != name_c]
    df_raw = df_raw[cols]
    return df_raw

def get_sessions():
    if DB:
        v = db_get("sessions")
        return v if v else DEFAULT_SESSIONS
    return M["sessions"]

def get_notes_label():
    if DB:
        v = db_get("notes_label")
        return v if v else "Données par défaut"
    return "Données par défaut"

def get_qr_codes():
    if DB:
        v = db_get("qr_codes")
        return v if (v and isinstance(v, dict)) else {}
    return M["qr_codes"]

def get_notes_visibles():
    if DB:
        v = db_get("notes_visibles")
        return bool(v) if v is not None else False
    return M["notes_visibles"]

def get_presences():
    if DB: return db_get_presences()
    if M["presences"]: return pd.DataFrame(M["presences"])
    return pd.DataFrame(columns=["agent","date_jour","heure","session","statut"])

def get_calendrier():
    if DB: return db_get_calendrier()
    if M["calendrier"]: return pd.DataFrame(M["calendrier"])
    return pd.DataFrame(columns=["date_jour","a_une_note","nom_note","ordre_note"])

def get_justifications():
    if DB: return db_get_justifications()
    if M["justifications"]: return pd.DataFrame(M["justifications"])
    return pd.DataFrame(columns=["agent","date_jour","session","type_absence","motif"])

def get_seuil_note():
    if DB:
        v = db_get_seuil("seuil_note_insuffisante")
        return float(v) if v else 12.0
    return 12.0

def get_heure_retard(session):
    is_matin = "matin" in session.lower()
    if is_matin:
        if DB:
            v = db_get_seuil("heure_retard_matin")
            return v.strip('"') if v else "08:00"
        return "08:00"
    else:
        if DB:
            v = db_get_seuil("heure_retard_apres_midi")
            return v.strip('"') if v else "14:45"
        return "14:45"

def save_notes(df, label=""):
    # Sauvegarder aussi le nom de la colonne des noms
    name_c = get_name_col(df)
    if DB:
        db_set("notes", df.to_dict("records"))
        db_set("notes_label", label)
        db_set("notes_name_col", name_c)
    else:
        M["notes"] = df.to_dict("records")
        M["notes_name_col"] = name_c

def save_sessions(lst):
    if DB: db_set("sessions", lst)
    else: M["sessions"] = lst

def save_qr_codes(qr_map):
    if DB: db_set("qr_codes", qr_map)
    else: M["qr_codes"] = qr_map

def set_notes_visibles(val):
    if DB: db_set("notes_visibles", val)
    else: M["notes_visibles"] = val
    db_get.clear()

def add_presence(agent, session, statut="Présent", date_jour=None):
    if date_jour is None: date_jour = datetime.date.today()
    now_time = datetime.datetime.now().strftime("%H:%M:%S")
    heure_limite = get_heure_retard(session)
    if now_time > heure_limite and statut == "Présent":
        statut = "En retard"
    if DB: return db_add_presence(agent, date_jour, now_time, session, statut)
    M["presences"].append({"agent":agent,"date_jour":date_jour,
        "heure":now_time,"session":session,"statut":statut})
    return True

def is_pointed(agent, session, date_jour=None):
    if date_jour is None: date_jour = datetime.date.today()
    if DB: return db_check_pointed(agent, session, date_jour)
    for p in M["presences"]:
        if p["agent"]==agent and p["session"]==session and p["date_jour"]==date_jour:
            return True, p["heure"]
    return False, None

def add_justification(agent, date_jour, session, type_abs, motif):
    if DB: return db_add_justification(agent, date_jour, session, type_abs, motif)
    M["justifications"].append({"agent":agent,"date_jour":date_jour,
        "session":session,"type_absence":type_abs,"motif":motif})
    return True

def load_gsheet(url):
    try:
        if "spreadsheets/d/" in url:
            sid = url.split("spreadsheets/d/")[1].split("/")[0]
            url = f"https://docs.google.com/spreadsheets/d/{sid}/export?format=csv"
        df = pd.read_csv(url)
        df.columns = df.columns.astype(str).str.strip()
        return df, None
    except Exception as e: return None, str(e)

def get_admin_pwd():
    try: return st.secrets["admin"]["password"]
    except: return "formation2026"

# ════════════════════════════════════════════════════════════
# QR CODES
# ════════════════════════════════════════════════════════════
QR_PREFIX = "AGENT::"
def agent_key(name): return f"{QR_PREFIX}{name}"
def key_to_agent(k): return k.split("::",1)[1] if "::" in k else k

def make_qr_b64(name, box_size=6):
    qr = qrcode.QRCode(version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_H,
        box_size=box_size, border=3)
    qr.add_data(agent_key(name)); qr.make(fit=True)
    img = qr.make_image(fill_color="#2c3e7a", back_color="white")
    buf = io.BytesIO(); img.save(buf, format="PNG"); buf.seek(0)
    return base64.b64encode(buf.getvalue()).decode()

def generate_and_store_qr(agents_list, box_size=6):
    qr_map = {}
    prog = st.progress(0, text="Génération des QR codes…")
    for i, agent in enumerate(agents_list):
        qr_map[agent] = make_qr_b64(agent, box_size)
        prog.progress((i+1)/len(agents_list), text=f"{i+1}/{len(agents_list)} — {agent[:28]}…")
    prog.empty()
    save_qr_codes(qr_map)
    return qr_map

# ════════════════════════════════════════════════════════════
# FONCTIONS MÉTIER
# ════════════════════════════════════════════════════════════
def mention_info(n, seuil=12):
    if n>=16:      return "Très Bien",  "badge-tb"
    elif n>=14:    return "Bien",        "badge-b"
    elif n>=seuil: return "Assez Bien", "badge-ab"
    else:          return "Insuffisant","badge-in"

def get_name_col(df):
    """La colonne des noms est toujours la première (get_notes() la met en premier)."""
    return df.columns[0]

def get_note_cols(df):
    """Toutes les colonnes sauf la première (qui est la colonne des noms)."""
    return list(df.columns[1:])

def compute_stats(df, note_cols, name_col, seuil=12):
    rows=[]
    for _,row in df.iterrows():
        vals=[float(v) if str(v).strip() not in ["","nan","None","NaN"] else 0.0 for v in [row[c] for c in note_cols]]
        moy=round(np.mean(vals),2); ml,_=mention_info(moy,seuil)
        rows.append({"Agent":row[name_col],"Moyenne":moy,"Max":max(vals),
                     "Min":min(vals),"Médiane":round(np.median(vals),2),
                     "Écart-type":round(np.std(vals),2),"Nb notes":len(vals),"Mention":ml})
    return pd.DataFrame(rows)

def compute_presence_stats(agents, pres_df, just_df, sessions_list):
    rows=[]
    for agent in agents:
        ap = pres_df[pres_df["agent"]==agent] if len(pres_df)>0 else pd.DataFrame()
        aj = just_df[just_df["agent"]==agent] if len(just_df)>0 else pd.DataFrame()
        total = len(sessions_list)
        nb_present  = len(ap[ap["statut"]=="Présent"])  if len(ap)>0 else 0
        nb_retard   = len(ap[ap["statut"]=="En retard"]) if len(ap)>0 else 0
        nb_pointe   = nb_present + nb_retard
        nb_absent   = total - nb_pointe
        nb_just     = len(aj) if len(aj)>0 else 0
        nb_abs_nj   = max(0, nb_absent - nb_just)
        taux        = round(nb_pointe/total*100,1) if total>0 else 0
        rows.append({
            "Agent":agent,"Présent":nb_present,"En retard":nb_retard,
            "Absent":nb_absent,"Absent justifié":nb_just,
            "Absent non justifié":nb_abs_nj,
            "Total sessions":total,"Taux présence (%)":taux,
            "Alerte":nb_abs_nj > 0,
        })
    return pd.DataFrame(rows)

def compute_profil(moy, taux, seuil_note=12, seuil_presence=75):
    bon_note = moy >= seuil_note
    bon_pres = taux >= seuil_presence
    if bon_note and bon_pres:       return "🟢 Performant et assidu",      "profil-vert"
    elif bon_note and not bon_pres: return "🟡 Performant mais absent",    "profil-bleu"
    elif not bon_note and bon_pres: return "🟠 Présent mais en difficulté","profil-orange"
    else:                           return "🔴 En difficulté et absent",   "profil-rouge"

def compute_cumul_absences_vs_notes(agent, note_cols, df_notes, df_cal, df_pres, name_col):
    rows=[]
    if len(df_cal)==0: return pd.DataFrame()
    agent_notes = df_notes[df_notes[name_col]==agent]
    if len(agent_notes)==0: return pd.DataFrame()
    agent_notes = agent_notes.iloc[0]
    notes_cal = df_cal[df_cal["a_une_note"]==True].sort_values("date_jour")
    agent_pres = df_pres[df_pres["agent"]==agent] if len(df_pres)>0 else pd.DataFrame()
    all_dates = sorted(df_cal["date_jour"].tolist())
    for _,cal_row in notes_cal.iterrows():
        nom_note = cal_row["nom_note"]
        date_note = cal_row["date_jour"]
        if nom_note not in note_cols: continue
        note_val = float(agent_notes[nom_note])
        dates_avant = [d for d in all_dates if d <= date_note]
        nb_abs = 0
        for d in dates_avant:
            if len(agent_pres)>0:
                pj = agent_pres[agent_pres["date_jour"]==d]
                if len(pj)==0: nb_abs += 2
                elif len(pj)<2: nb_abs += 1
            else: nb_abs += 2
        rows.append({"Note":nom_note,"Date":date_note,"Valeur note":note_val,"Absences cumulées":nb_abs})
    return pd.DataFrame(rows)

def get_cat_color(n, seuil=12):
    if n>=16:      return "#e6f4ea","🟢"
    elif n>=14:    return "#dbeafe","🔵"
    elif n>=seuil: return "#fef3c7","🟡"
    else:          return "#fee2e2","🔴"

# ════════════════════════════════════════════════════════════
# EXPORTS
# ════════════════════════════════════════════════════════════
def export_couleur_excel(df, note_cols, name_col, df_stats, seuil=12):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    buf = io.BytesIO(); WB = Workbook(); ws = WB.active; ws.title = "Notes couleur"
    COLORS = {"tb":("1e7e34","c6efce"),"b":("1a56db","bdd7ee"),"ab":("92400e","ffeb9c"),"in":("b91c1c","ffc7ce")}
    ARROWS = {"tb":"▲","b":"→","ab":"→","in":"▼"}
    def get_cat(n):
        if n>=16: return "tb"
        elif n>=14: return "b"
        elif n>=seuil: return "ab"
        else: return "in"
    thin = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
    headers=[name_col]+note_cols+["Moyenne","Mention"]
    for ci,h in enumerate(headers,1):
        c=ws.cell(row=1,column=ci,value=h)
        c.font=Font(bold=True,color="FFFFFF"); c.fill=PatternFill("solid",fgColor="2c3e7a")
        c.alignment=Alignment(horizontal="center",vertical="center"); c.border=thin
    for ri,row in df.iterrows():
        r=ri+2; ws.cell(row=r,column=1,value=row[name_col]).border=thin
        for ci,col in enumerate(note_cols,2):
            v=float(row[col]); cat=get_cat(v); fc,bg=COLORS[cat]
            c=ws.cell(row=r,column=ci,value=f"{ARROWS[cat]} {int(v)}")
            c.fill=PatternFill("solid",fgColor=bg); c.font=Font(color=fc,bold=True)
            c.alignment=Alignment(horizontal="center"); c.border=thin
        moy_v=df_stats[df_stats["Agent"]==row[name_col]]["Moyenne"].values
        moy_val=moy_v[0] if len(moy_v)>0 else 0; cat=get_cat(moy_val); fc,bg=COLORS[cat]
        cm=ws.cell(row=r,column=len(note_cols)+2,value=f"{ARROWS[cat]} {moy_val:.2f}")
        cm.fill=PatternFill("solid",fgColor=bg); cm.font=Font(color=fc,bold=True)
        cm.alignment=Alignment(horizontal="center"); cm.border=thin
        ml,_=mention_info(moy_val,seuil); ws.cell(row=r,column=len(note_cols)+3,value=ml).border=thin
    ws.column_dimensions["A"].width=32
    for ci in range(2,len(headers)+1): ws.column_dimensions[get_column_letter(ci)].width=14
    lr=len(df)+4; ws.cell(row=lr,column=1,value="LÉGENDE").font=Font(bold=True)
    for i,(lbl,cat) in enumerate([("▲ Très Bien (≥16)","tb"),("→ Bien (14-15)","b"),
                                    (f"→ Assez Bien ({seuil}-13)","ab"),(f"▼ Insuffisant (<{seuil})","in")]):
        fc,bg=COLORS[cat]; c=ws.cell(row=lr+1+i,column=1,value=lbl)
        c.fill=PatternFill("solid",fgColor=bg); c.font=Font(color=fc,bold=True)
    WB.save(buf); buf.seek(0); return buf.getvalue()

def export_stats_excel(df, note_cols, name_col, df_stats, df_pres_stats):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    buf = io.BytesIO(); WB = Workbook()
    thin = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
    ws1=WB.active; ws1.title="Notes individuelles"
    h1=[name_col]+note_cols+["Moyenne","Médiane","Écart-type","Mention"]
    for ci,h in enumerate(h1,1):
        c=ws1.cell(row=1,column=ci,value=h)
        c.font=Font(bold=True,color="FFFFFF"); c.fill=PatternFill("solid",fgColor="4a6cf7")
        c.alignment=Alignment(horizontal="center"); c.border=thin
    for ri,row in df.iterrows():
        r=ri+2; ws1.cell(row=r,column=1,value=row[name_col]).border=thin
        for ci,col in enumerate(note_cols,2):
            c=ws1.cell(row=r,column=ci,value=float(row[col]))
            c.alignment=Alignment(horizontal="center"); c.border=thin
        s=df_stats[df_stats["Agent"]==row[name_col]]
        if len(s)>0:
            s=s.iloc[0]
            for ci2,val in enumerate([s["Moyenne"],s["Médiane"],s["Écart-type"],s["Mention"]],len(note_cols)+2):
                c=ws1.cell(row=r,column=ci2,value=val)
                c.alignment=Alignment(horizontal="center"); c.border=thin
                if ci2==len(note_cols)+2:
                    v=s["Moyenne"]
                    bg=("c6efce" if v>=16 else "bdd7ee" if v>=14 else "ffeb9c" if v>=12 else "ffc7ce")
                    c.fill=PatternFill("solid",fgColor=bg); c.font=Font(bold=True)
    ws1.column_dimensions["A"].width=32
    for ci in range(2,len(h1)+1): ws1.column_dimensions[get_column_letter(ci)].width=13
    ws1.freeze_panes="B2"
    if df_pres_stats is not None and len(df_pres_stats)>0:
        ws2=WB.create_sheet("Présences"); h2=list(df_pres_stats.columns)
        for ci,h in enumerate(h2,1):
            c=ws2.cell(row=1,column=ci,value=h)
            c.font=Font(bold=True,color="FFFFFF"); c.fill=PatternFill("solid",fgColor="2c3e7a")
            c.alignment=Alignment(horizontal="center"); c.border=thin
        for ri,row in df_pres_stats.iterrows():
            r=ri+2
            for ci,col in enumerate(h2,1):
                v=row[col]; c=ws2.cell(row=r,column=ci,value=v)
                c.alignment=Alignment(horizontal="center"); c.border=thin
                if col=="Absent non justifié" and v>0:
                    c.fill=PatternFill("solid",fgColor="ffc7ce"); c.font=Font(color="b91c1c",bold=True)
                if col=="Alerte": c.value="⚠️ OUI" if v else "✅ OK"
        ws2.column_dimensions["A"].width=32
        for ci in range(2,len(h2)+1): ws2.column_dimensions[get_column_letter(ci)].width=16
        ws3=WB.create_sheet("Profils agents")
        df_croise=df_stats.merge(df_pres_stats[["Agent","Taux présence (%)","Absent non justifié","Alerte"]],
                                  on="Agent",how="left").fillna(0)
        df_croise["Profil"]=df_croise.apply(lambda r: compute_profil(r["Moyenne"],r["Taux présence (%)"])[0],axis=1)
        h3=["Agent","Moyenne","Mention","Taux présence (%)","Absent non justifié","Alerte","Profil"]
        for ci,h in enumerate(h3,1):
            c=ws3.cell(row=1,column=ci,value=h)
            c.font=Font(bold=True,color="FFFFFF"); c.fill=PatternFill("solid",fgColor="2c3e7a")
            c.alignment=Alignment(horizontal="center"); c.border=thin
        PROFIL_COLORS={"🟢":"c6efce","🟡":"ffeb9c","🟠":"ffe0b2","🔴":"ffc7ce"}
        for ri,row in df_croise.iterrows():
            r=ri+2
            for ci,col in enumerate(h3,1):
                v=row[col]; c=ws3.cell(row=r,column=ci,value=v)
                c.alignment=Alignment(horizontal="center"); c.border=thin
                if col=="Alerte": c.value="⚠️ OUI" if v else "✅ OK"
                if col=="Profil":
                    bg=PROFIL_COLORS.get(str(v)[:2],"FFFFFF")
                    c.fill=PatternFill("solid",fgColor=bg)
        ws3.column_dimensions["A"].width=32
        for ci in range(2,len(h3)+1): ws3.column_dimensions[get_column_letter(ci)].width=20
    ws4=WB.create_sheet("Synthèse")
    all_v=df[note_cols].values.flatten().astype(float)
    stats_g=[
        ("Nombre d'agents",len(df)),("Nombre d'évaluations",len(note_cols)),
        ("Moyenne générale",round(np.mean(all_v),2)),
        ("Note maximum",int(df[note_cols].max().max())),("Note minimum",int(df[note_cols].min().min())),
        ("Très Bien (≥16)",len(df_stats[df_stats["Mention"]=="Très Bien"])),
        ("Bien (14-15)",len(df_stats[df_stats["Mention"]=="Bien"])),
        ("Assez Bien",len(df_stats[df_stats["Mention"]=="Assez Bien"])),
        ("Insuffisant",len(df_stats[df_stats["Mention"]=="Insuffisant"])),
        ("Notes 20/20",int((df[note_cols]==20).sum().sum())),
        ("Notes = 0",int((df[note_cols]==0).sum().sum())),
    ]
    for ci,h in enumerate(["Indicateur","Valeur"],1):
        c=ws4.cell(row=1,column=ci,value=h)
        c.font=Font(bold=True,color="FFFFFF"); c.fill=PatternFill("solid",fgColor="2c3e7a")
        c.alignment=Alignment(horizontal="center"); c.border=thin
    for ri,(k,v) in enumerate(stats_g,2):
        ws4.cell(row=ri,column=1,value=k).border=thin; ws4.cell(row=ri,column=2,value=v).border=thin
    ws4.column_dimensions["A"].width=28; ws4.column_dimensions["B"].width=16
    WB.save(buf); buf.seek(0); return buf.getvalue()

def sp(text):
    text = str(text)
    return unicodedata.normalize('NFKD', text).encode('latin-1','ignore').decode('latin-1')

def export_pdf_direction(df_stats, df_pres_stats, note_cols, agents, seuil_note=12):
    if not PDF_OK: return None
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica","B",20); pdf.set_text_color(44,62,122)
    pdf.cell(0,12,sp("RAPPORT DE FORMATION"),ln=True,align="C")
    pdf.set_font("Helvetica","",11); pdf.set_text_color(107,125,179)
    pdf.cell(0,8,sp(f"Genere le {datetime.date.today().strftime('%d/%m/%Y')}"),ln=True,align="C")
    pdf.ln(6)
    pdf.set_font("Helvetica","B",14); pdf.set_text_color(44,62,122)
    pdf.cell(0,10,sp("1. Statistiques generales"),ln=True)
    pdf.set_draw_color(74,108,247); pdf.line(10,pdf.get_y(),200,pdf.get_y()); pdf.ln(4)
    all_v=df_stats["Moyenne"].values
    pdf.set_font("Helvetica","",11); pdf.set_text_color(26,26,46)
    stats=[("Nombre d agents",len(agents)),("Nombre evaluations",len(note_cols)),
           ("Moyenne generale",round(np.mean(all_v),2)),
           ("Tres Bien (>=16)",len(df_stats[df_stats["Mention"]=="Très Bien"])),
           ("Bien (14-15)",len(df_stats[df_stats["Mention"]=="Bien"])),
           ("Assez Bien",len(df_stats[df_stats["Mention"]=="Assez Bien"])),
           ("Insuffisant",len(df_stats[df_stats["Mention"]=="Insuffisant"]))]
    for k,v in stats:
        pdf.cell(100,8,sp(k),border="B"); pdf.cell(90,8,sp(str(v)),border="B",ln=True,align="C")
    pdf.ln(8)
    pdf.set_font("Helvetica","B",14); pdf.set_text_color(44,62,122)
    pdf.cell(0,10,sp("2. Top 5 agents"),ln=True)
    pdf.set_draw_color(74,108,247); pdf.line(10,pdf.get_y(),200,pdf.get_y()); pdf.ln(4)
    top5=df_stats.sort_values("Moyenne",ascending=False).head(5)
    pdf.set_font("Helvetica","B",10); pdf.set_fill_color(44,62,122); pdf.set_text_color(255,255,255)
    pdf.cell(10,8,"#",fill=True,border=1,align="C")
    pdf.cell(110,8,sp("Agent"),fill=True,border=1)
    pdf.cell(40,8,sp("Moyenne"),fill=True,border=1,align="C")
    pdf.cell(30,8,sp("Mention"),fill=True,border=1,align="C",ln=True)
    pdf.set_font("Helvetica","",10); pdf.set_text_color(26,26,46)
    for i,(_,row) in enumerate(top5.iterrows(),1):
        pdf.set_fill_color(240,244,255) if i%2==0 else pdf.set_fill_color(255,255,255)
        pdf.cell(10,8,str(i),fill=True,border=1,align="C")
        pdf.cell(110,8,sp(str(row["Agent"])[:45]),fill=True,border=1)
        pdf.cell(40,8,sp(f"{row['Moyenne']:.2f}/20"),fill=True,border=1,align="C")
        pdf.cell(30,8,sp(str(row["Mention"])),fill=True,border=1,align="C",ln=True)
    pdf.ln(8)
    if df_pres_stats is not None and len(df_pres_stats)>0:
        alertes=df_pres_stats[df_pres_stats["Alerte"]==True]
        pdf.set_font("Helvetica","B",14); pdf.set_text_color(44,62,122)
        pdf.cell(0,10,sp(f"3. Alertes - {len(alertes)} agent(s) en situation critique"),ln=True)
        pdf.set_draw_color(231,76,60); pdf.line(10,pdf.get_y(),200,pdf.get_y()); pdf.ln(4)
        if len(alertes)>0:
            pdf.set_font("Helvetica","B",10); pdf.set_fill_color(44,62,122); pdf.set_text_color(255,255,255)
            pdf.cell(110,8,sp("Agent"),fill=True,border=1)
            pdf.cell(40,8,sp("Abs. non justifiee"),fill=True,border=1,align="C")
            pdf.cell(40,8,sp("Taux presence"),fill=True,border=1,align="C",ln=True)
            pdf.set_font("Helvetica","",10); pdf.set_text_color(26,26,46)
            for _,row in alertes.iterrows():
                pdf.set_fill_color(255,235,235)
                pdf.cell(110,8,sp(str(row["Agent"])[:45]),fill=True,border=1)
                pdf.cell(40,8,sp(str(row["Absent non justifié"])),fill=True,border=1,align="C")
                pdf.cell(40,8,sp(f"{row['Taux présence (%)']:.1f}%"),fill=True,border=1,align="C",ln=True)
        else:
            pdf.set_font("Helvetica","",11); pdf.set_text_color(30,126,52)
            pdf.cell(0,8,sp("Aucune alerte - tous les agents sont dans les normes."),ln=True)
    pdf.ln(8); pdf.set_font("Helvetica","I",9); pdf.set_text_color(160,174,192)
    pdf.cell(0,6,sp("Rapport genere automatiquement par la Plateforme Formation"),align="C")
    return bytes(pdf.output())

# ════════════════════════════════════════════════════════════
# DÉTECTION VUE
# ════════════════════════════════════════════════════════════
try:
    _vue_param = st.query_params.get("vue", "")
    if isinstance(_vue_param, list):
        _vue_param = _vue_param[0] if _vue_param else ""
    VUE_PUBLIC = str(_vue_param).strip() == "public"
except Exception:
    VUE_PUBLIC = False
# ════════════════════════════════════════════════════════════
# SESSION STATE
# ════════════════════════════════════════════════════════════
sessions = get_sessions()
for k,v in {
    "admin_ok":False,"camera_on":False,"camera_facing":"environment",
    "last_agent":None,"last_time":None,
    "session_active": sessions[0] if sessions else "Session 1",
}.items():
    if k not in st.session_state: st.session_state[k]=v
if st.session_state.session_active not in sessions:
    st.session_state.session_active = sessions[0]

# ════════════════════════════════════════════════════════════
# CHARGEMENT
# ════════════════════════════════════════════════════════════
df         = get_notes()
name_col   = get_name_col(df)          # détecte automatiquement la colonne des noms
note_cols  = get_note_cols(df)         # détecte automatiquement les colonnes numériques
agents     = df[name_col].tolist()
seuil_note = get_seuil_note()
df_stats   = compute_stats(df, note_cols, name_col, seuil_note)
all_vals   = df[note_cols].values.flatten().astype(float)
pres_df    = get_presences()
just_df    = get_justifications()
cal_df     = get_calendrier()
label      = get_notes_label()
notes_visibles = get_notes_visibles()
df_pres_stats  = compute_presence_stats(agents, pres_df, just_df, sessions)

# ════════════════════════════════════════════════════════════
# SIDEBAR
# ════════════════════════════════════════════════════════════
with st.sidebar:
    if DB:
        st.markdown('<div class="db-ok">🟢 <b>Supabase connecté</b><br>'
                    '<small>Données persistantes · synchro auto</small></div>',
                    unsafe_allow_html=True)
    else:
        st.markdown('<div class="db-warn">⚠️ <b>Mode mémoire</b><br>'
                    '<small>Configurez Supabase dans secrets.toml</small></div>',
                    unsafe_allow_html=True)
    st.markdown(f"**Source :** {label}")
    st.markdown(f"**{len(agents)}** agents · **{len(note_cols)}** évaluations")
    if VUE_PUBLIC:
        st.info("👁️ Vue publique")
    st.markdown("---")
    st.markdown("### 🗓️ Session")
    st.session_state.session_active = st.selectbox(
        "Session active", sessions,
        index=sessions.index(st.session_state.session_active)
              if st.session_state.session_active in sessions else 0,
        label_visibility="collapsed")
    st.markdown("---")
    if not VUE_PUBLIC:
        st.markdown("### 🔍 Filtres")
        sel_agents = st.multiselect("Agents (courbes)", options=agents, default=agents[:5])
        note_range = st.slider("Filtre moyenne", 0.0, 20.0, (0.0,20.0), 0.5)
        st.markdown("---")
    nb_pts  = len(pres_df)
    nb_sess = len(pres_df[pres_df["session"]==st.session_state.session_active]) if nb_pts>0 else 0
    st.metric("Total pointages", nb_pts)
    st.metric("Présents session", nb_sess)
    if not VUE_PUBLIC:
        nb_alertes = len(df_pres_stats[df_pres_stats["Alerte"]==True])
        if nb_alertes > 0:
            st.error(f"⚠️ {nb_alertes} alerte(s) absence")
        if DB and st.button("🔄 Actualiser"):
            db_get.clear(); db_get_presences.clear()
            db_get_justifications.clear(); db_get_calendrier.clear()
            st.rerun()

# ════════════════════════════════════════════════════════════
# EN-TÊTE
# ════════════════════════════════════════════════════════════
if VUE_PUBLIC:
    st.markdown('<div class="main-title">📊 Résultats de Formation</div>', unsafe_allow_html=True)
    st.markdown(f'<div class="sub-title">{len(agents)} participants · {len(note_cols)} évaluations</div>',
                unsafe_allow_html=True)
else:
    st.markdown('<div class="main-title">📊 Plateforme Formation — Direction</div>', unsafe_allow_html=True)
    st.markdown(f'<div class="sub-title">{len(agents)} agents · {len(note_cols)} évaluations · '
                f'<b>{label}</b> · Session : <b>{st.session_state.session_active}</b></div>',
                unsafe_allow_html=True)

    # KPI direction uniquement
    moy_gen   = round(np.mean(all_vals),2)
    moy_max_a = round(df[note_cols].astype(float).max(axis=1).mean(),2)
    moy_min_a = round(df[note_cols].astype(float).min(axis=1).mean(),2)
    nb_20     = int((df[note_cols]==20).sum().sum())
    nb_zero   = int((df[note_cols]==0).sum().sum())
    r1 = st.columns(5)
    for col,(val,lbl,sub) in zip(r1,[
        (moy_gen,"Moyenne Générale","toutes notes"),
        (moy_max_a,"Moy. Max/Agent","meilleur score moyen"),
        (moy_min_a,"Moy. Min/Agent","score le plus bas moyen"),
        (nb_20,"Notes 20/20","scores parfaits"),
        (nb_zero,"Notes = 0","zéros obtenus"),
    ]):
        col.markdown(f'<div class="kpi-card"><div class="kpi-value">{val}</div>'
                     f'<div class="kpi-label">{lbl}</div><div class="kpi-sub">{sub}</div></div>',
                     unsafe_allow_html=True)
    st.markdown("<br>",unsafe_allow_html=True)
    r2 = st.columns(4)
    for col,(val,lbl,sub) in zip(r2,[
        (len(df_stats[df_stats["Mention"]=="Très Bien"]),"Très Bien",f"sur {len(agents)} agents"),
        (len(df_stats[df_stats["Mention"]=="Bien"]),"Bien",f"sur {len(agents)} agents"),
        (len(df_stats[df_stats["Mention"]=="Assez Bien"]),"Assez Bien",f"sur {len(agents)} agents"),
        (len(df_stats[df_stats["Mention"]=="Insuffisant"]),"Insuffisant",f"sur {len(agents)} agents"),
    ]):
        col.markdown(f'<div class="kpi-card"><div class="kpi-value">{val}</div>'
                     f'<div class="kpi-label">{lbl}</div><div class="kpi-sub">{sub}</div></div>',
                     unsafe_allow_html=True)
    st.markdown("<br>",unsafe_allow_html=True)

# ════════════════════════════════════════════════════════════
# ONGLETS
# ════════════════════════════════════════════════════════════

if VUE_PUBLIC:
    # ──────────────────────────────────────────────────────────
    # VUE PUBLIQUE
    # ──────────────────────────────────────────────────────────
    tab_lbl_chiffres = "📊 Notes" if notes_visibles else "🔒 Notes (masquées)"
    tab_lbl_podium   = "🏆 Classement" if notes_visibles else "🔒 Classement (masqué)"

    tab_couleur, tab_fleches, tab_chiffres, tab_podium_p, tab_pres_p, tab_kiosque_p = st.tabs([
        "🎨 Notes couleur",
        "🏹 Notes flèches",
        tab_lbl_chiffres,
        tab_lbl_podium,
        "👥 Présences",
        "📷 Scanner",
    ])

    # ── Helper flèche + couleur ────────────────────────────────
    def get_arrow_info(n, seuil=12):
        if n >= 16:   return "↑", "#e6f4ea", "#1e7e34"
        elif n >= 14: return "↗", "#dbeafe", "#1a56db"
        elif n >= seuil: return "→", "#fef3c7", "#92400e"
        else:         return "↓", "#fee2e2", "#b91c1c"

    # ── Notes couleur (toujours visible) ──────────────────────
    with tab_couleur:
        st.markdown('<div class="section-title">🎨 Résultats par couleur</div>',
                    unsafe_allow_html=True)
        st.caption("Les chiffres sont masqués — seul le niveau est visible.")

        html_t = '<div style="overflow-x:auto"><table style="width:100%;border-collapse:collapse;font-size:.82rem;">'
        html_t += '<tr style="background:#2c3e7a;color:white;">'
        html_t += f'<th style="padding:8px;text-align:left;border:1px solid #dce4f5;min-width:160px;">{name_col}</th>'
        for col in note_cols:
            html_t += f'<th style="padding:8px;text-align:center;border:1px solid #dce4f5;">{col}</th>'
        html_t += '<th style="padding:8px;text-align:center;border:1px solid #dce4f5;">Bilan</th>'
        html_t += '</tr>'
        for _, row in df.iterrows():
            html_t += '<tr>'
            html_t += (f'<td style="padding:7px 10px;border:1px solid #f0f4ff;'
                       f'font-weight:500;white-space:nowrap;">{row[name_col]}</td>')
            for col in note_cols:
                v = float(row[col])
                bg, emoji = get_cat_color(v, seuil_note)
                html_t += (f'<td style="padding:7px;text-align:center;'
                           f'background:{bg};border:1px solid #f0f4ff;font-size:1rem;">{emoji}</td>')
            moy_v = df_stats[df_stats["Agent"]==row[name_col]]["Moyenne"].values
            if len(moy_v)>0:
                bg,emoji = get_cat_color(moy_v[0], seuil_note)
                html_t += (f'<td style="padding:7px;text-align:center;background:{bg};'
                           f'border:1px solid #f0f4ff;font-weight:700;font-size:1rem;">{emoji}</td>')
            html_t += '</tr>'
        html_t += '</table></div>'
        st.markdown(html_t, unsafe_allow_html=True)
        st.markdown("<br>",unsafe_allow_html=True)

        l1,l2,l3,l4 = st.columns(4)
        for col,emoji,lbl,bg in [
            (l1,"🟢",f"Très Bien (≥16)","#e6f4ea"),
            (l2,"🔵","Bien (14-15)","#dbeafe"),
            (l3,"🟡",f"Assez Bien ({int(seuil_note)}-13)","#fef3c7"),
            (l4,"🔴",f"Insuffisant (<{int(seuil_note)})","#fee2e2"),
        ]:
            col.markdown(f'<div style="background:{bg};border-radius:8px;padding:10px;'
                         f'text-align:center;font-size:.85rem;font-weight:500;">{emoji} {lbl}</div>',
                         unsafe_allow_html=True)

    # ── Notes flèches ──────────────────────────────────────────
    with tab_fleches:
        st.markdown('<div class="section-title">🏹 Notes par flèches</div>',unsafe_allow_html=True)

        # Barre de recherche
        search_query = st.text_input(
            "🔍 Rechercher un agent",
            placeholder="Tapez un nom...",
            key="search_fleches"
        )

        # Filtrer les agents selon la recherche
        df_filtered = df.copy()
        if search_query.strip():
            mask = df[name_col].astype(str).str.lower().str.contains(
                search_query.strip().lower(), na=False)
            df_filtered = df[mask].reset_index(drop=True)

        if len(df_filtered) == 0:
            st.warning(f"Aucun agent trouvé pour **{search_query}**")
        else:
            # Tableau flèches
            html_f = '<div style="overflow-x:auto"><table style="width:100%;border-collapse:collapse;font-size:.82rem;">'
            html_f += '<tr style="background:#2c3e7a;color:white;">'
            html_f += f'<th style="padding:8px;text-align:left;border:1px solid #dce4f5;min-width:160px;">{name_col}</th>'
            for col in note_cols:
                html_f += f'<th style="padding:8px;text-align:center;border:1px solid #dce4f5;">{col}</th>'
            html_f += '<th style="padding:8px;text-align:center;border:1px solid #dce4f5;">Bilan</th>'
            html_f += '</tr>'

            for _, row in df_filtered.iterrows():
                html_f += '<tr>'
                html_f += (f'<td style="padding:7px 10px;border:1px solid #f0f4ff;'
                           f'font-weight:500;white-space:nowrap;">{row[name_col]}</td>')
                for col in note_cols:
                    v = float(row[col])
                    arrow, bg, fc = get_arrow_info(v, seuil_note)
                    if notes_visibles:
                        content = (f'<span style="font-size:1.1rem;">{arrow}</span>'
                                   f'<span style="font-size:.85rem;margin-left:4px;">{int(v)}</span>')
                    else:
                        content = f'<span style="font-size:1.3rem;">{arrow}</span>'
                    html_f += (f'<td style="padding:7px;text-align:center;background:{bg};'
                               f'color:{fc};border:1px solid #f0f4ff;font-weight:600;">{content}</td>')
                # Bilan
                moy_v = df_stats[df_stats["Agent"]==row[name_col]]["Moyenne"].values
                if len(moy_v)>0:
                    arrow, bg, fc = get_arrow_info(moy_v[0], seuil_note)
                    if notes_visibles:
                        content_b = (f'<span style="font-size:1.1rem;">{arrow}</span>'
                                     f'<span style="font-size:.85rem;margin-left:4px;">{moy_v[0]:.1f}</span>')
                    else:
                        content_b = f'<span style="font-size:1.3rem;">{arrow}</span>'
                    html_f += (f'<td style="padding:7px;text-align:center;background:{bg};'
                               f'color:{fc};border:1px solid #f0f4ff;font-weight:700;">{content_b}</td>')
                html_f += '</tr>'
            html_f += '</table></div>'
            st.markdown(html_f, unsafe_allow_html=True)

        st.markdown("<br>",unsafe_allow_html=True)
        # Légende flèches
        l1,l2,l3,l4 = st.columns(4)
        for col,arrow,lbl,bg,fc in [
            (l1,"↑",f"Très Bien (≥16)","#e6f4ea","#1e7e34"),
            (l2,"↗","Bien (14-15)","#dbeafe","#1a56db"),
            (l3,"→",f"Assez Bien ({int(seuil_note)}-13)","#fef3c7","#92400e"),
            (l4,"↓",f"Insuffisant (<{int(seuil_note)})","#fee2e2","#b91c1c"),
        ]:
            col.markdown(
                f'<div style="background:{bg};border-radius:8px;padding:10px;text-align:center;'
                f'font-size:.85rem;font-weight:600;color:{fc};">'
                f'<span style="font-size:1.2rem;">{arrow}</span> {lbl}</div>',
                unsafe_allow_html=True)
        if not notes_visibles:
            st.caption("🔒 Les chiffres sont masqués — activables par l'administrateur.")

    # ── Notes chiffres ─────────────────────────────────────────
    with tab_chiffres:
        if not notes_visibles:
            st.markdown('<div class="locked-box">'
                        '<div style="font-size:3rem;">🔒</div>'
                        '<div style="font-size:1.2rem;color:#6b7db3;margin-top:12px;font-weight:500;">'
                        'Notes détaillées non disponibles</div>'
                        '<div style="font-size:.9rem;color:#a0aec0;margin-top:8px;">'
                        "L'administrateur les activera prochainement.</div></div>",
                        unsafe_allow_html=True)
        else:
            st.markdown('<div class="section-title">📊 Notes détaillées</div>', unsafe_allow_html=True)
            heat = df.set_index(name_col)[note_cols].astype(float)
            fig5 = go.Figure(go.Heatmap(
                z=heat.values, x=note_cols, y=heat.index.tolist(),
                colorscale=[[0,"#fee2e2"],[0.001,"#fecaca"],[0.3,"#bfdbfe"],[0.6,"#4a6cf7"],[1,"#1a237e"]],
                text=heat.values.astype(int), texttemplate="%{text}",
                textfont=dict(size=10),
                hovertemplate="<b>%{y}</b><br>%{x} : <b>%{z}</b>/20<extra></extra>",
                colorbar=dict(title="Note"), zmin=0, zmax=20))
            fig5.update_layout(paper_bgcolor="white",plot_bgcolor="white",
                font=dict(family="DM Sans",color="#2c3e7a"),
                xaxis=dict(side="top"),yaxis=dict(autorange="reversed"),
                height=max(400,len(agents)*28),margin=dict(t=50,b=20,l=240,r=20))
            st.plotly_chart(fig5, use_container_width=True)

    # ── Classement ─────────────────────────────────────────────
    with tab_podium_p:
        if not notes_visibles:
            st.markdown('<div class="locked-box">'
                        '<div style="font-size:3rem;">🔒</div>'
                        '<div style="font-size:1.2rem;color:#6b7db3;margin-top:12px;font-weight:500;">'
                        'Classement non disponible</div>'
                        '<div style="font-size:.9rem;color:#a0aec0;margin-top:8px;">'
                        "L'administrateur l'activera prochainement.</div></div>",
                        unsafe_allow_html=True)
        else:
            st.markdown('<div class="section-title">🏆 Classement général</div>', unsafe_allow_html=True)
            drk = df_stats.sort_values("Moyenne",ascending=False).reset_index(drop=True)
            drk.index += 1
            if len(drk)>=3:
                p2,p1,p3 = st.columns([1,1.2,1])
                for col,rank,medal,bg in [(p1,1,"🥇","#FFF9C4"),(p2,2,"🥈","#F5F5F5"),(p3,3,"🥉","#FBE9E7")]:
                    r=drk.iloc[rank-1]; ml,mc=mention_info(r["Moyenne"],seuil_note)
                    col.markdown(f'<div style="background:{bg};border-radius:16px;padding:16px;'
                                 f'text-align:center;box-shadow:0 4px 16px rgba(0,0,0,.07);">'
                                 f'<div style="font-size:2rem;">{medal}</div>'
                                 f'<div style="font-weight:700;color:#2c3e7a;font-size:.9rem;">{r["Agent"]}</div>'
                                 f'<div style="font-size:1.8rem;font-weight:700;color:#4a6cf7;">{r["Moyenne"]:.2f}</div>'
                                 f'<span class="{mc}">{ml}</span></div>',unsafe_allow_html=True)
            st.markdown("<br>",unsafe_allow_html=True)
            cb_ = ["#2d6a4f" if m>=16 else "#1a56db" if m>=14 else "#92400e" if m>=seuil_note else "#b91c1c"
                   for m in drk["Moyenne"]]
            fb = go.Figure(go.Bar(x=drk["Moyenne"],y=drk["Agent"],orientation="h",
                marker=dict(color=cb_,line=dict(color="white",width=0.5)),
                text=[f"{v:.2f}" for v in drk["Moyenne"]],textposition="outside"))
            fb.update_layout(paper_bgcolor="white",plot_bgcolor="white",
                font=dict(family="DM Sans",color="#2c3e7a"),
                xaxis=dict(range=[0,23],gridcolor="#f0f4ff"),
                yaxis=dict(autorange="reversed",tickfont=dict(size=10)),
                height=max(400,len(agents)*28),margin=dict(t=20,b=40,l=230,r=60),showlegend=False)
            st.plotly_chart(fb, use_container_width=True)

    # ── Présences publique ─────────────────────────────────────
    with tab_pres_p:
        st.markdown('<div class="section-title">👥 Présences</div>', unsafe_allow_html=True)
        pv = get_presences()
        if len(pv)==0:
            st.info("Aucun pointage enregistré pour l'instant.")
        else:
            dp = compute_presence_stats(agents, pv, get_justifications(), sessions)
            
            nb_p = len(dp[(dp["Présent"]+dp["En retard"])>0])
            taux_m = round(dp[(dp["Présent"]+dp["En retard"])>0]["Taux présence (%)"].mean(),1) if nb_p>0 else 0
            pa,pb,pc = st.columns(3)
            for col,(val,lbl,sub) in zip([pa,pb,pc],[
                (f"{taux_m}%","Taux moyen","agents pointés"),
                (nb_p,"Agents pointés","au moins 1 session"),
                (len(pv),"Total pointages","enregistrés"),
            ]):
                col.markdown(f'<div class="kpi-card"><div class="kpi-value">{val}</div>'
                             f'<div class="kpi-label">{lbl}</div><div class="kpi-sub">{sub}</div></div>',
                             unsafe_allow_html=True)
            st.markdown("<br>",unsafe_allow_html=True)
            ds = dp[(dp["Présent"]+dp["En retard"])>0].sort_values("Taux présence (%)",ascending=True)
            
            if len(ds)>0:
                cp = ["#27ae60" if t>=75 else "#f39c12" if t>=50 else "#e74c3c" for t in ds["Taux présence (%)"]]
                fp = go.Figure(go.Bar(x=ds["Taux présence (%)"],y=ds["Agent"],orientation="h",
                    marker=dict(color=cp,line=dict(color="white",width=0.5)),
                    text=[f"{v}%" for v in ds["Taux présence (%)"]],textposition="outside"))
                fp.update_layout(paper_bgcolor="white",plot_bgcolor="white",
                    font=dict(family="DM Sans",color="#2c3e7a"),
                    xaxis=dict(range=[0,118],gridcolor="#f0f4ff"),
                    yaxis=dict(autorange="reversed",tickfont=dict(size=10)),
                    height=max(360,len(ds)*30),margin=dict(t=20,b=40,l=230,r=60),showlegend=False)
                st.plotly_chart(fp, use_container_width=True)

    # ── Kiosque scanner public ─────────────────────────────────
    with tab_kiosque_p:
        st.markdown('<div class="section-title">📷 Scanner QR — Pointage</div>',unsafe_allow_html=True)
        st.markdown(f"**Session :** `{st.session_state.session_active}`")
        st.components.v1.html("""
        <script>
        async function wl(){try{if('wakeLock' in navigator)await navigator.wakeLock.request('screen');}catch(e){}}
        wl();document.addEventListener('visibilitychange',()=>{if(document.visibilityState==='visible')wl();});
        </script>
        <div style="background:#f0fdf4;border:1px solid #86efac;border-radius:8px;
                    padding:8px 14px;font-size:13px;color:#166534;text-align:center;">
            📱 Écran maintenu actif (Wake Lock Android)
        </div>""", height=50)
        ck1,ck2 = st.columns([1.3,1])
        with ck1:
            # Sélecteur de caméra (toujours visible)
            cam_choice = st.radio(
                "📷 Caméra",
                options=["Arrière (recommandé)", "Avant"],
                index=0 if st.session_state.camera_facing=="environment" else 1,
                horizontal=True, key="cam_pub"
            )
            st.session_state.camera_facing = "environment" if "Arrière" in cam_choice else "user"

            if not st.session_state.camera_on:
                if st.button("📷 Ouvrir la caméra",type="primary",use_container_width=True):
                    st.session_state.camera_on=True; st.session_state.last_agent=None; st.rerun()
            else:
                if st.button("🔴 Fermer la caméra",type="secondary",use_container_width=True):
                    st.session_state.camera_on=False; st.rerun()
                now_str = datetime.datetime.now().strftime("%H:%M:%S")
                st.markdown(f'<div style="background:#1a1a2e;border-radius:12px;padding:14px;'
                            f'text-align:center;margin:8px 0;">'
                            f'<div style="font-size:2.2rem;font-weight:700;color:#4a6cf7;'
                            f'font-family:monospace;">{now_str}</div>'
                            f'<div style="font-size:.8rem;color:#6b7db3;margin-top:4px;">'
                            f'Retard matin après 08:00 · Après-midi après 14:45</div></div>',
                            unsafe_allow_html=True)
                if st.session_state.last_agent:
                    st.markdown(f'<div class="scan-ok">✅ <b>{st.session_state.last_agent}</b><br>'
                                f'<small>{st.session_state.last_time} · Prêt pour le suivant…</small></div>',
                                unsafe_allow_html=True)
                if QR_OK:
                    facing = st.session_state.camera_facing  # "environment" ou "user"
                    st.caption(f"📸 Caméra {'arrière' if facing=='environment' else 'avant'} — Présentez le badge QR")
                    # Essai avec camera_facing_mode, puis facing_mode, puis sans argument
                    # On change la key selon la caméra pour forcer la réinitialisation du composant
                    scanner_key = f"pub_scanner_{facing}"
                    qr_val = None
                    try:
                        qr_val = qrcode_scanner(key=scanner_key, camera_facing_mode=facing)
                    except TypeError:
                        try:
                            qr_val = qrcode_scanner(key=scanner_key, facing_mode=facing)
                        except TypeError:
                            # Le module ne supporte aucun paramètre de caméra :
                            # on injecte un script JS pour forcer le facingMode côté navigateur
                            force_js = f"""
<script>
(function waitForVideo(){{
  var v = document.querySelector('video');
  if(!v){{ setTimeout(waitForVideo, 300); return; }}
  if(v._facingPatched) return;
  v._facingPatched = true;
  var origGetUserMedia = navigator.mediaDevices.getUserMedia.bind(navigator.mediaDevices);
  navigator.mediaDevices.getUserMedia = function(c){{
    if(c && c.video && typeof c.video === 'object'){{
      c.video.facingMode = {{ ideal: '{facing}' }};
    }} else if(c && c.video === true){{
      c.video = {{ facingMode: {{ ideal: '{facing}' }} }};
    }}
    return origGetUserMedia(c);
  }};
  v.srcObject && v.srcObject.getTracks().forEach(t=>t.stop());
}})();
</script>"""
                            st.components.v1.html(force_js, height=0)
                            qr_val = qrcode_scanner(key=scanner_key)
                    if qr_val:
                        af = key_to_agent(str(qr_val).strip())
                        if af in agents:
                            deja,h = is_pointed(af, st.session_state.session_active)
                            if not deja:
                                add_presence(af, st.session_state.session_active)
                                st.session_state.last_agent = af
                                st.session_state.last_time  = datetime.datetime.now().strftime("%H:%M:%S")
                                st.rerun()
                            else: st.warning(f"ℹ️ **{af}** déjà pointé à {h}")
                        else: st.error(f"❌ QR non reconnu : `{af}`")
                else:
                    st.error("❌ Module scanner non disponible.")
        with ck2:
            st.markdown("**Pointage manuel**")
            with st.form("fm_pub"):
                ag = st.selectbox("Agent", agents)
                st_ = st.selectbox("Statut",["Présent","En retard","Excusé"])
                if st.form_submit_button("✅ Pointer",use_container_width=True):
                    deja,h = is_pointed(ag, st.session_state.session_active)
                    if not deja:
                        add_presence(ag, st.session_state.session_active, st_)
                        st.success(f"✅ {ag}"); st.rerun()
                    else: st.info(f"Déjà pointé à {h}")
            today = datetime.date.today()
            pj = pres_df[pres_df["date_jour"]==today] if len(pres_df)>0 else pd.DataFrame()
            pj_s = pj[pj["session"]==st.session_state.session_active] if len(pj)>0 else pd.DataFrame()
            if len(pj_s)>0:
                st.dataframe(pj_s[["agent","heure","statut"]].sort_values("heure",ascending=False),
                             use_container_width=True,hide_index=True,height=260)

else:
    # ──────────────────────────────────────────────────────────
    # VUE DIRECTION — tous les onglets
    # ──────────────────────────────────────────────────────────
    sel_agents = sel_agents if 'sel_agents' in dir() else agents[:5]
    note_range = note_range if 'note_range' in dir() else (0.0, 20.0)

    (tab_pres, tab_alertes, tab_analyse, tab_croise,
     tab_notes, tab_heatmap, tab_podium, tab_export,
     tab_kiosque, tab_admin) = st.tabs([
        "👥 Présences","⚠️ Alertes","📊 Analyse cumulative",
        "🔗 Notes × Présences","📈 Courbes notes","🔥 Heatmap",
        "🏆 Classement","📤 Exports","📷 Kiosque","🔐 Admin",
    ])

    # ── PRÉSENCES ─────────────────────────────────────────────
    with tab_pres:
        st.markdown('<div class="section-title">👥 Présences & Ponctualité</div>',unsafe_allow_html=True)
        if len(pres_df)==0:
            st.info("Aucun pointage. Utilisez le **📷 Kiosque** ou pointez dans **🔐 Admin**.")
        else:
            pa,pb,pc,pd_,pe = st.columns(5)
            for col,(val,lbl,sub) in zip([pa,pb,pc,pd_,pe],[
                (int(df_pres_stats["Présent"].sum()),"Présences","total"),
                (int(df_pres_stats["En retard"].sum()),"Retards","total"),
                (int(df_pres_stats["Absent"].sum()),"Absences","total"),
                (int(df_pres_stats["Absent justifié"].sum()),"Abs. justifiées","avec motif"),
                (int(df_pres_stats["Absent non justifié"].sum()),"Abs. non justifiées","⚠️ alertes"),
            ]):
                col.markdown(f'<div class="kpi-card"><div class="kpi-value">{val}</div>'
                             f'<div class="kpi-label">{lbl}</div><div class="kpi-sub">{sub}</div></div>',
                             unsafe_allow_html=True)
            st.markdown("<br>",unsafe_allow_html=True)
            ds = df_pres_stats.sort_values("Taux présence (%)",ascending=True)
            cp = ["#27ae60" if t>=75 else "#f39c12" if t>=50 else "#e74c3c" for t in ds["Taux présence (%)"]]
            fp = go.Figure(go.Bar(x=ds["Taux présence (%)"],y=ds["Agent"],orientation="h",
                marker=dict(color=cp,line=dict(color="white",width=0.5)),
                text=[f"{v}%" for v in ds["Taux présence (%)"]],textposition="outside"))
            fp.update_layout(paper_bgcolor="white",plot_bgcolor="white",
                font=dict(family="DM Sans",color="#2c3e7a"),
                xaxis=dict(range=[0,118],gridcolor="#f0f4ff"),
                yaxis=dict(autorange="reversed",tickfont=dict(size=10)),
                height=max(360,len(ds)*30),margin=dict(t=20,b=40,l=230,r=60),showlegend=False,
                shapes=[dict(type="line",xref="x",x0=75,x1=75,yref="paper",y0=0,y1=1,
                             line=dict(color="#27ae60",width=1.5,dash="dot"))])
            st.plotly_chart(fp,use_container_width=True)
            def spr(row):
                return [("background-color:#fee2e2;color:#b91c1c" if row[c]>0 else "") if c=="Absent non justifié"
                        else ("background-color:#fef3c7;color:#92400e" if row[c]>0 else "") if c=="En retard"
                        else "" for c in row.index]
            df_show = df_pres_stats.copy()
            df_show["Alerte"] = df_show["Alerte"].map({True:"⚠️ OUI",False:"✅ OK"})
            st.dataframe(df_show.style.apply(spr,axis=1).format({"Taux présence (%)":"{:.1f}%"}),
                         use_container_width=True,height=420)
            st.markdown('<div class="section-title">✏️ Justifier une absence / retard</div>',
                        unsafe_allow_html=True)
            with st.form("form_just"):
                cj1,cj2,cj3 = st.columns(3)
                with cj1: ag_j = st.selectbox("Agent",agents)
                with cj2: date_j = st.date_input("Date",value=datetime.date.today())
                with cj3: sess_j = st.selectbox("Session",sessions)
                cj4,cj5 = st.columns(2)
                with cj4: type_j = st.selectbox("Type",["Absence","Retard"])
                with cj5: motif_j = st.selectbox("Motif",MOTIFS)
                if st.form_submit_button("✅ Enregistrer",use_container_width=True):
                    add_justification(ag_j,date_j,sess_j,type_j,motif_j)
                    st.success(f"✅ Justification enregistrée pour {ag_j}"); st.rerun()

    # ── ALERTES ───────────────────────────────────────────────
    with tab_alertes:
        st.markdown('<div class="section-title">⚠️ Alertes</div>',unsafe_allow_html=True)
        alertes_df = df_pres_stats[df_pres_stats["Alerte"]==True]
        insuf_df   = df_stats[df_stats["Mention"]=="Insuffisant"]
        ca,cb = st.columns(2)
        with ca:
            st.markdown(f"**{len(alertes_df)} agent(s) — absence non justifiée :**")
            if len(alertes_df)>0:
                for _,row in alertes_df.iterrows():
                    st.markdown(f'<div class="alerte-rouge">⚠️ <b>{row["Agent"]}</b> — '
                                f'{row["Absent non justifié"]} abs. · '
                                f'Taux : {row["Taux présence (%)"]:.1f}%</div>',unsafe_allow_html=True)
            else: st.success("✅ Aucune absence non justifiée")
        with cb:
            st.markdown(f"**{len(insuf_df)} agent(s) sous le seuil ({seuil_note}/20) :**")
            if len(insuf_df)>0:
                for _,row in insuf_df.iterrows():
                    st.markdown(f'<div class="alerte-orange">📉 <b>{row["Agent"]}</b> — '
                                f'{row["Moyenne"]:.2f}/20</div>',unsafe_allow_html=True)
            else: st.success("✅ Tous au-dessus du seuil")
        st.markdown('<div class="section-title">🎯 Profils combinés</div>',unsafe_allow_html=True)
        if len(df_pres_stats)>0:
            df_profils = df_stats.merge(df_pres_stats[["Agent","Taux présence (%)"]],on="Agent",how="left").fillna(0)
            df_profils["Profil"] = df_profils.apply(
                lambda r: compute_profil(r["Moyenne"],r["Taux présence (%)"],seuil_note)[0], axis=1)
            for profil,group in df_profils.groupby("Profil"):
                emoji = str(profil)[:2]
                css = {"🟢":"profil-vert","🟡":"profil-bleu","🟠":"profil-orange","🔴":"profil-rouge"}.get(emoji,"profil-vert")
                agents_str = " · ".join([str(a) for a in group["Agent"].tolist()])
                st.markdown(f'<div class="{css}"><b>{profil}</b> ({len(group)} agents)<br>'
                            f'<small>{agents_str}</small></div>',unsafe_allow_html=True)

    # ── ANALYSE CUMULATIVE ────────────────────────────────────
    with tab_analyse:
        st.markdown('<div class="section-title">📊 Absences cumulées → Impact sur notes</div>',
                    unsafe_allow_html=True)
        if len(cal_df)==0:
            st.info("Configurez le calendrier dans **🔐 Admin → 🗓️ Calendrier**.")
        else:
            agent_sel_a = st.selectbox("Agent", agents, key="sel_analyse")
            df_cumul = compute_cumul_absences_vs_notes(agent_sel_a,note_cols,df,cal_df,pres_df,name_col)
            if len(df_cumul)>0:
                fig_c = go.Figure()
                fig_c.add_trace(go.Scatter(x=df_cumul["Note"],y=df_cumul["Valeur note"],
                    mode="lines+markers+text",name="Note",
                    line=dict(color="#4a6cf7",width=3,shape="spline"),
                    marker=dict(size=10,color="#4a6cf7",line=dict(color="white",width=2)),
                    text=[str(int(v)) for v in df_cumul["Valeur note"]],
                    textposition="top center",textfont=dict(size=10,color="#2c3e7a"),yaxis="y1"))
                fig_c.add_trace(go.Bar(x=df_cumul["Note"],y=df_cumul["Absences cumulées"],
                    name="Absences cumulées",opacity=0.4,marker=dict(color="#e74c3c"),yaxis="y2"))
                fig_c.update_layout(paper_bgcolor="white",plot_bgcolor="white",
                    font=dict(family="DM Sans",color="#2c3e7a"),
                    xaxis=dict(title="Évaluation",gridcolor="#f0f4ff"),
                    yaxis=dict(title="Note /20",range=[0,21],gridcolor="#f0f4ff",side="left"),
                    yaxis2=dict(title="Absences cumulées",overlaying="y",side="right",showgrid=False),
                    hovermode="x unified",height=420,margin=dict(t=30,b=40,l=50,r=60))
                st.plotly_chart(fig_c,use_container_width=True)
            else:
                st.warning("Pas assez de données. Vérifiez le calendrier.")

    # ── CROISEMENT ────────────────────────────────────────────
    with tab_croise:
        st.markdown('<div class="section-title">🔗 Notes × Présences</div>',unsafe_allow_html=True)
        if len(pres_df)==0:
            st.info("Aucune présence enregistrée.")
        else:
            dcx = df_stats.merge(df_pres_stats[["Agent","Taux présence (%)","Absent non justifié"]],
                                  on="Agent",how="left").fillna(0)
            dcx["Statut"] = dcx.apply(
                lambda r: compute_profil(r["Moyenne"],r["Taux présence (%)"],seuil_note)[0],axis=1)
            cm = {"🟢 Performant et assidu":"#27ae60","🟡 Performant mais absent":"#f39c12",
                  "🟠 Présent mais en difficulté":"#f97316","🔴 En difficulté et absent":"#e74c3c"}
            fsc = go.Figure()
            for sv,color in cm.items():
                sub = dcx[dcx["Statut"]==sv]
                if len(sub)>0:
                    fsc.add_trace(go.Scatter(x=sub["Taux présence (%)"],y=sub["Moyenne"],
                        mode="markers+text",name=sv,
                        marker=dict(size=12,color=color,opacity=0.85,line=dict(color="white",width=1.5)),
                        text=sub["Agent"].str.split().str[0],textposition="top center",textfont=dict(size=9),
                        hovertemplate="<b>%{text}</b><br>%{x}% · %{y:.2f}<extra></extra>"))
            pts = dcx[dcx["Taux présence (%)"]>0]
            if len(pts)>3:
                try:
                    z=np.polyfit(pts["Taux présence (%)"],pts["Moyenne"],1); p=np.poly1d(z)
                    xl=np.linspace(0,100,50)
                    fsc.add_trace(go.Scatter(x=xl,y=p(xl),mode="lines",name="Tendance",
                        line=dict(color="#4a6cf7",width=2,dash="dash"),hoverinfo="skip"))
                    corr=round(pts["Taux présence (%)"].corr(pts["Moyenne"]),3)
                    st.info(f"**Corrélation présence / notes : {corr}**")
                except: pass
            fsc.update_layout(paper_bgcolor="white",plot_bgcolor="white",
                font=dict(family="DM Sans",color="#2c3e7a"),
                xaxis=dict(title="Taux présence (%)",range=[-5,110],gridcolor="#f0f4ff"),
                yaxis=dict(title="Moyenne /20",range=[0,21],gridcolor="#f0f4ff"),
                height=480,margin=dict(t=40,b=40,l=50,r=20))
            st.plotly_chart(fsc,use_container_width=True)

    # ── COURBES NOTES ─────────────────────────────────────────
    with tab_notes:
        st.markdown('<div class="section-title">📈 Évolution des notes</div>',unsafe_allow_html=True)
        if not sel_agents:
            st.info("👈 Sélectionnez des agents dans le panneau latéral.")
        else:
            pal = (px.colors.qualitative.Pastel+px.colors.qualitative.Safe
                   +px.colors.qualitative.Vivid+px.colors.qualitative.Dark24)
            fig = go.Figure()
            for i,agent in enumerate(sel_agents):
                row = df[df[name_col]==agent].iloc[0]
                notes = [float(row[c]) for c in note_cols]; c = pal[i%len(pal)]
                mc = []
                for n in notes:
                    if n>=16: mc.append("#27ae60")
                    elif n>=14: mc.append("#1a56db")
                    elif n>=seuil_note: mc.append("#f39c12")
                    else: mc.append("#e74c3c")
                fig.add_trace(go.Scatter(x=note_cols,y=notes,mode="lines+markers+text",name=agent,
                    line=dict(color=c,width=2,shape="spline"),
                    marker=dict(size=11,color=mc,line=dict(color="white",width=1.5)),
                    text=[str(int(n)) for n in notes],textposition="top center",textfont=dict(size=9,color=c),
                    hovertemplate=f"<b>{agent}</b><br>%{{x}} : <b>%{{y}}</b>/20<extra></extra>"))
            fig.update_layout(paper_bgcolor="white",plot_bgcolor="white",
                font=dict(family="DM Sans",color="#2c3e7a"),
                xaxis=dict(title="Évaluation",gridcolor="#f0f4ff"),
                yaxis=dict(title="Note /20",range=[-0.5,21],gridcolor="#f0f4ff"),
                legend=dict(bgcolor="rgba(255,255,255,.95)",bordercolor="#dce4f5",borderwidth=1),
                hovermode="x unified",height=480,margin=dict(t=30,b=40,l=50,r=80),
                shapes=[dict(type="line",xref="paper",x0=0,x1=1,yref="y",y0=seuil_note,y1=seuil_note,
                             line=dict(color="#e74c3c",width=1.5,dash="dot")),
                        dict(type="line",xref="paper",x0=0,x1=1,yref="y",y0=14,y1=14,
                             line=dict(color="#27ae60",width=1.5,dash="dot"))],
                annotations=[dict(xref="paper",x=1.01,yref="y",y=seuil_note,
                                   text=f"Seuil {int(seuil_note)}",showarrow=False,
                                   font=dict(size=9,color="#e74c3c"),xanchor="left"),
                             dict(xref="paper",x=1.01,yref="y",y=14,text="Seuil 14",
                                   showarrow=False,font=dict(size=9,color="#27ae60"),xanchor="left")])
            st.plotly_chart(fig,use_container_width=True)

    # ── HEATMAP ───────────────────────────────────────────────
    with tab_heatmap:
        st.markdown('<div class="section-title">🔥 Carte de chaleur</div>',unsafe_allow_html=True)
        heat = df.set_index(name_col)[note_cols].astype(float)
        f5 = go.Figure(go.Heatmap(z=heat.values,x=note_cols,y=heat.index.tolist(),
            colorscale=[[0,"#fee2e2"],[0.001,"#fecaca"],[0.3,"#bfdbfe"],[0.6,"#4a6cf7"],[1,"#1a237e"]],
            text=heat.values.astype(int),texttemplate="%{text}",textfont=dict(size=10),
            hovertemplate="<b>%{y}</b><br>%{x} : <b>%{z}</b>/20<extra></extra>",
            colorbar=dict(title="Note"),zmin=0,zmax=20))
        f5.update_layout(paper_bgcolor="white",plot_bgcolor="white",
            font=dict(family="DM Sans",color="#2c3e7a"),
            xaxis=dict(side="top"),yaxis=dict(autorange="reversed"),
            height=max(400,len(agents)*28),margin=dict(t=50,b=20,l=240,r=20))
        st.plotly_chart(f5,use_container_width=True)

    # ── CLASSEMENT ────────────────────────────────────────────
    with tab_podium:
        st.markdown('<div class="section-title">🏆 Classement</div>',unsafe_allow_html=True)
        drk = df_stats.sort_values("Moyenne",ascending=False).reset_index(drop=True); drk.index+=1
        if len(drk)>=3:
            p2,p1,p3 = st.columns([1,1.2,1])
            for col,rank,medal,bg in [(p1,1,"🥇","#FFF9C4"),(p2,2,"🥈","#F5F5F5"),(p3,3,"🥉","#FBE9E7")]:
                r=drk.iloc[rank-1]; ml,mc=mention_info(r["Moyenne"],seuil_note)
                col.markdown(f'<div style="background:{bg};border-radius:16px;padding:16px;'
                             f'text-align:center;box-shadow:0 4px 16px rgba(0,0,0,.07);">'
                             f'<div style="font-size:2rem;">{medal}</div>'
                             f'<div style="font-weight:700;color:#2c3e7a;font-size:.9rem;">{r["Agent"]}</div>'
                             f'<div style="font-size:1.8rem;font-weight:700;color:#4a6cf7;">{r["Moyenne"]:.2f}</div>'
                             f'<span class="{mc}">{ml}</span></div>',unsafe_allow_html=True)
        st.markdown("<br>",unsafe_allow_html=True)
        cb_ = ["#2d6a4f" if m>=16 else "#1a56db" if m>=14 else "#92400e" if m>=seuil_note else "#b91c1c"
               for m in drk["Moyenne"]]
        fb = go.Figure(go.Bar(x=drk["Moyenne"],y=drk["Agent"],orientation="h",
            marker=dict(color=cb_,line=dict(color="white",width=0.5)),
            text=[f"{v:.2f}" for v in drk["Moyenne"]],textposition="outside"))
        fb.update_layout(paper_bgcolor="white",plot_bgcolor="white",
            font=dict(family="DM Sans",color="#2c3e7a"),
            xaxis=dict(range=[0,23],gridcolor="#f0f4ff"),
            yaxis=dict(autorange="reversed",tickfont=dict(size=10)),
            height=max(400,len(agents)*28),margin=dict(t=20,b=40,l=230,r=60),showlegend=False)
        st.plotly_chart(fb,use_container_width=True)

    # ── EXPORTS ───────────────────────────────────────────────
    with tab_export:
        st.markdown('<div class="section-title">📤 Exports</div>',unsafe_allow_html=True)
        e1,e2 = st.columns(2)
        with e1:
            st.markdown("#### 🎨 Export Couleur + Flèches")
            st.markdown("Notes remplacées par couleurs et flèches ▲→▼")
            if st.button("⬇️ Générer export couleur",type="primary",use_container_width=True):
                data = export_couleur_excel(df,note_cols,name_col,df_stats,seuil_note)
                st.download_button("💾 Télécharger — Notes couleur",data=data,
                    file_name=f"notes_couleur_{datetime.date.today()}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True)
        with e2:
            st.markdown("#### 📊 Export Stats + Présences + Profils")
            st.markdown("4 feuilles : notes · présences · profils · synthèse")
            if st.button("⬇️ Générer export stats",type="primary",use_container_width=True):
                data = export_stats_excel(df,note_cols,name_col,df_stats,df_pres_stats)
                st.download_button("💾 Télécharger — Stats complètes",data=data,
                    file_name=f"stats_formation_{datetime.date.today()}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True)
        st.markdown("---")
        st.markdown("#### 📄 Rapport PDF Direction")
        if st.button("⬇️ Générer PDF",type="primary",use_container_width=True):
            if PDF_OK:
                data = export_pdf_direction(df_stats,df_pres_stats,note_cols,agents,seuil_note)
                if data:
                    st.download_button("💾 Télécharger PDF",data=data,
                        file_name=f"rapport_direction_{datetime.date.today()}.pdf",
                        mime="application/pdf",use_container_width=True)
            else: st.error("❌ Module fpdf2 manquant.")

    # ── KIOSQUE ───────────────────────────────────────────────
    with tab_kiosque:
        st.markdown('<div class="section-title">📷 Mode Kiosque — Tablette Android</div>',
                    unsafe_allow_html=True)
        st.markdown(f"**Session :** `{st.session_state.session_active}`")
        st.components.v1.html("""
        <script>
        async function wl(){try{if('wakeLock' in navigator)await navigator.wakeLock.request('screen');}catch(e){}}
        wl();document.addEventListener('visibilitychange',()=>{if(document.visibilityState==='visible')wl();});
        </script>
        <div style="background:#f0fdf4;border:1px solid #86efac;border-radius:8px;
                    padding:8px 14px;font-size:13px;color:#166534;text-align:center;">
            📱 Écran maintenu actif (Wake Lock activé)
        </div>""", height=50)
        ck1,ck2 = st.columns([1.3,1])
        with ck1:
            # Sélecteur de caméra
            cam_choice_dir = st.radio(
                "📷 Caméra",
                options=["Arrière (recommandé)", "Avant"],
                index=0 if st.session_state.camera_facing=="environment" else 1,
                horizontal=True, key="cam_dir"
            )
            st.session_state.camera_facing = "environment" if "Arrière" in cam_choice_dir else "user"

            if not st.session_state.camera_on:
                if st.button("📷 Ouvrir la caméra",type="primary",use_container_width=True):
                    st.session_state.camera_on=True; st.session_state.last_agent=None; st.rerun()
            else:
                if st.button("🔴 Fermer la caméra",type="secondary",use_container_width=True):
                    st.session_state.camera_on=False; st.rerun()
                now_str = datetime.datetime.now().strftime("%H:%M:%S")
                st.markdown(f'<div style="background:#1a1a2e;border-radius:12px;padding:16px;'
                            f'text-align:center;margin:8px 0;">'
                            f'<div style="font-size:2.5rem;font-weight:700;color:#4a6cf7;'
                            f'font-family:monospace;">{now_str}</div>'
                            f'<div style="font-size:.9rem;color:#6b7db3;margin-top:4px;">'
                            f'Retard matin après 08:00 · Après-midi après 14:45</div></div>',
                            unsafe_allow_html=True)
                if st.session_state.last_agent:
                    st.markdown(f'<div class="scan-ok" style="font-size:1.1rem;">'
                                f'✅ <b>{st.session_state.last_agent}</b><br>'
                                f'<small>Pointé à {st.session_state.last_time} · '
                                f'Prêt pour le prochain…</small></div>',unsafe_allow_html=True)
                if QR_OK:
                    st.caption(f"📸 Caméra {'arrière' if st.session_state.camera_facing=='environment' else 'avant'} — Présentez le badge QR")
                    qr_val = qrcode_scanner(key="kiosque_dir", camera_facing_mode=st.session_state.camera_facing)
                    if qr_val:
                        af = key_to_agent(str(qr_val).strip())
                        if af in agents:
                            deja,h = is_pointed(af,st.session_state.session_active)
                            if not deja:
                                add_presence(af,st.session_state.session_active)
                                st.session_state.last_agent = af
                                st.session_state.last_time  = datetime.datetime.now().strftime("%H:%M:%S")
                                st.rerun()
                            else: st.warning(f"ℹ️ **{af}** déjà pointé à {h}")
                        else: st.error(f"❌ QR non reconnu : `{af}`")
                else: st.error("❌ Module scanner non disponible.")
        with ck2:
            st.markdown("**Pointage manuel rapide**")
            with st.form("fm_kiosque"):
                ag = st.selectbox("Agent",agents)
                st_ = st.selectbox("Statut",["Présent","En retard","Excusé"])
                if st.form_submit_button("✅ Pointer",use_container_width=True):
                    deja,h = is_pointed(ag,st.session_state.session_active)
                    if not deja:
                        add_presence(ag,st.session_state.session_active,st_)
                        st.success(f"✅ {ag}"); st.rerun()
                    else: st.info(f"Déjà pointé à {h}")
            today = datetime.date.today()
            pj = pres_df[pres_df["date_jour"]==today] if len(pres_df)>0 else pd.DataFrame()
            pj_s = pj[pj["session"]==st.session_state.session_active] if len(pj)>0 else pd.DataFrame()
            if len(pj_s)>0:
                st.dataframe(pj_s[["agent","heure","statut"]].sort_values("heure",ascending=False),
                             use_container_width=True,hide_index=True,height=300)
            else: st.info("Aucun pointage aujourd'hui.")

    # ── ADMIN ─────────────────────────────────────────────────
    with tab_admin:
        st.markdown('<div class="section-title">🔐 Administration</div>',unsafe_allow_html=True)
        if not st.session_state.admin_ok:
            pwd = st.text_input("Mot de passe",type="password")
            if st.button("🔓 Se connecter",type="primary",use_container_width=True):
                if pwd==get_admin_pwd():
                    st.session_state.admin_ok=True; st.rerun()
                else: st.error("❌ Mot de passe incorrect")
            st.caption("Par défaut : `formation2026`")
        else:
            ca,cb_ = st.columns([2,1])
            with ca: st.success(f"✅ Admin · {'🟢 Supabase' if DB else '⚠️ Mémoire'}")
            with cb_:
                if st.button("🚪 Déconnexion",use_container_width=True):
                    st.session_state.admin_ok=False; st.rerun()

            a1,a2,a3,a4,a5 = st.tabs(["📊 Notes","🎫 QR Codes","🗓️ Calendrier","⚙️ Seuils & Vue","📋 Données"])

            # ── Notes ──────────────────────────────────────────
            with a1:
                st.markdown('<div class="section-title">Importer les notes</div>',unsafe_allow_html=True)
                src = st.radio("Source",["📁 Fichier Excel","🔗 Google Sheets"],horizontal=True)
                if "Excel" in src:
                    fn = st.file_uploader("Fichier Excel",type=["xlsx","xls"])
                    if fn and st.button("✅ Publier pour tous",type="primary",use_container_width=True):
                        try:
                            df_new=pd.read_excel(fn); df_new.columns=df_new.columns.astype(str).str.strip()
                            save_notes(df_new,f"Excel — {df_new.shape[0]} agents")
                            st.success(f"✅ {df_new.shape[0]} agents publiés !"); st.rerun()
                        except Exception as e: st.error(str(e))
                else:
                    url_gs = st.text_input("URL Google Sheets")
                    if st.button("✅ Publier",type="primary",use_container_width=True):
                        if url_gs:
                            df_new,err = load_gsheet(url_gs)
                            if err: st.error(err)
                            else:
                                save_notes(df_new,f"Google Sheets — {df_new.shape[0]} agents")
                                st.success(f"✅ {df_new.shape[0]} agents !"); st.rerun()
                if st.button("🔄 Données par défaut",type="secondary"):
                    if DB:
                        try:
                            sb.table("app_config").delete().in_("cle",["notes","notes_label"]).execute()
                            db_get.clear()
                        except: pass
                    else: M["notes"]=None
                    st.success("Rétabli."); st.rerun()

            # ── QR Codes ───────────────────────────────────────
            with a2:
                st.markdown('<div class="section-title">QR Codes — générés une fois</div>',unsafe_allow_html=True)
                existing_qr = get_qr_codes()
                if len(existing_qr)>0:
                    st.success(f"✅ {len(existing_qr)} QR codes en base — permanents.")
                    if st.button("🔄 Régénérer",type="secondary",use_container_width=True):
                        generate_and_store_qr(agents); st.rerun()
                else:
                    st.warning("⚠️ Aucun QR code. Générez-les une fois.")
                    if st.button(f"🔄 Générer {len(agents)} QR codes",type="primary",use_container_width=True):
                        generate_and_store_qr(agents); st.rerun()
                qr_map = get_qr_codes()
                if qr_map:
                    ags_dl = st.multiselect("Agents (vide=tous)",list(qr_map.keys()))
                    to_dl  = ags_dl if ags_dl else list(qr_map.keys())
                    if st.button(f"⬇️ ZIP ({len(to_dl)} badges)",use_container_width=True):
                        zb = io.BytesIO()
                        with zipfile.ZipFile(zb,"w",zipfile.ZIP_DEFLATED) as zf:
                            for ag,b64 in qr_map.items():
                                if ag in to_dl:
                                    zf.writestr(f"QR_Badges/{ag[:60].replace('/','_')}.png",base64.b64decode(b64))
                        zb.seek(0)
                        st.download_button("💾 Télécharger",data=zb.getvalue(),
                            file_name="QR_Badges.zip",mime="application/zip",use_container_width=True)
                    hg = '<div class="qr-grid">'
                    for ag,b64 in list(qr_map.items())[:12]:
                        hg += (f'<div class="qr-item"><img src="data:image/png;base64,{b64}" '
                               f'style="width:100%;border-radius:5px">'
                               f'<div class="qr-name">{ag[:30]}</div></div>')
                    st.markdown(hg+'</div>',unsafe_allow_html=True)

            # ── Calendrier ─────────────────────────────────────
            with a3:
                st.markdown('<div class="section-title">Calendrier — Dates et notes</div>',unsafe_allow_html=True)
                date_debut = st.date_input("Date de début",value=datetime.date.today())
                nb_jours   = st.number_input("Nombre de jours",min_value=1,max_value=90,value=21)
                if st.button("🗓️ Générer le calendrier",type="primary",use_container_width=True):
                    rows = []
                    for j in range(nb_jours):
                        d = date_debut+datetime.timedelta(days=j)
                        if d.weekday()<5:
                            rows.append({"date_jour":str(d),"a_une_note":False,"nom_note":None,"ordre_note":None})
                    if DB: db_save_calendrier(rows)
                    else: M["calendrier"]=[{**r,"date_jour":datetime.date.fromisoformat(r["date_jour"])} for r in rows]
                    st.success(f"✅ {len(rows)} jours générés."); st.rerun()
                cal = get_calendrier()
                if len(cal)>0:
                    st.markdown("**Associez chaque jour à une note :**")
                    cal_edit = cal.copy()
                    for i,row in cal.iterrows():
                        c1,c2,c3 = st.columns([2,1,2])
                        with c1: st.write(str(row["date_jour"]))
                        with c2: has_note = st.checkbox("Note ?",value=bool(row["a_une_note"]),key=f"cn_{i}")
                        with c3:
                            if has_note:
                                nn = st.selectbox("Note",["—"]+note_cols,key=f"cs_{i}",
                                    index=note_cols.index(row["nom_note"])+1 if row["nom_note"] in note_cols else 0)
                                cal_edit.at[i,"a_une_note"]=True
                                cal_edit.at[i,"nom_note"]=nn if nn!="—" else None
                            else:
                                cal_edit.at[i,"a_une_note"]=False; cal_edit.at[i,"nom_note"]=None
                    if st.button("💾 Sauvegarder",type="primary",use_container_width=True):
                        rows_s = []
                        for i,row in cal_edit.iterrows():
                            rows_s.append({"date_jour":str(row["date_jour"]),"a_une_note":bool(row["a_une_note"]),
                                           "nom_note":row["nom_note"],"ordre_note":int(i)})
                        if DB: db_save_calendrier(rows_s)
                        else: M["calendrier"]=cal_edit.to_dict("records")
                        st.success("✅ Calendrier sauvegardé !"); st.rerun()

            # ── Seuils & Vue ────────────────────────────────────
            with a4:
                # ── Activation notes vue publique ──────────────
                st.markdown('<div class="section-title">👁️ Visibilité — Vue publique</div>',
                            unsafe_allow_html=True)
                st.markdown("""
La **vue publique** (`?vue=public`) montre les notes en couleur par défaut.
Activez ici pour rendre les chiffres et le classement visibles à tous.
                """)
                col_vis1,col_vis2 = st.columns([2,1])
                with col_vis1:
                    if notes_visibles:
                        st.success("✅ Notes **visibles** — chiffres et classement affichés")
                    else:
                        st.warning("🔒 Notes **masquées** — couleurs seulement")
                with col_vis2:
                    if notes_visibles:
                        if st.button("🔒 Masquer les notes",use_container_width=True):
                            set_notes_visibles(False); st.rerun()
                    else:
                        if st.button("✅ Activer les notes",type="primary",use_container_width=True):
                            set_notes_visibles(True); st.rerun()

                st.markdown("**Liens à partager :**")
                base_url = "https://votre-app.streamlit.app"
                st.code(f"Vue direction (complète) : {base_url}\nVue publique (partageable) : {base_url}?vue=public")

                st.markdown("---")
                st.markdown('<div class="section-title">⚙️ Seuils et horaires</div>',unsafe_allow_html=True)
                s1,s2,s3 = st.columns(3)
                with s1: new_seuil=st.number_input("Seuil insuffisant",min_value=1,max_value=20,value=int(seuil_note))
                with s2: new_hm=st.text_input("Retard matin","08:00")
                with s3: new_ha=st.text_input("Retard après-midi","14:45")
                if st.button("💾 Sauvegarder seuils",type="primary",use_container_width=True):
                    if DB:
                        db_set_seuil("seuil_note_insuffisante",str(new_seuil))
                        db_set_seuil("heure_retard_matin",f'"{new_hm}"')
                        db_set_seuil("heure_retard_apres_midi",f'"{new_ha}"')
                    st.success("✅ Seuils sauvegardés."); st.rerun()

                st.markdown("---")
                st.markdown('<div class="section-title">Sessions</div>',unsafe_allow_html=True)
                txt = st.text_area("Sessions (1 par ligne)",value="\n".join(get_sessions()),height=160)
                if st.button("💾 Sauvegarder sessions",use_container_width=True):
                    ns=[s.strip() for s in txt.split("\n") if s.strip()]
                    if ns: save_sessions(ns); st.success(f"✅ {len(ns)} sessions."); st.rerun()

            # ── Données brutes ──────────────────────────────────
            with a5:
                st.markdown('<div class="section-title">Pointages & Justifications</div>',unsafe_allow_html=True)
                t_p,t_j = st.tabs(["Pointages","Justifications"])
                with t_p:
                    pa_ = get_presences()
                    if len(pa_)>0:
                        st.dataframe(pa_,use_container_width=True,height=260)
                        bc = io.BytesIO(); pa_.to_csv(bc,index=False,encoding="utf-8-sig")
                        st.download_button("⬇️ CSV",bc.getvalue(),file_name="pointages.csv",mime="text/csv")
                        if st.button("🗑️ Effacer tous",type="secondary"):
                            if DB: sb.table("presences").delete().neq("id",0).execute(); db_get_presences.clear()
                            else: M["presences"].clear()
                            st.success("Effacé."); st.rerun()
                    else: st.info("Aucun pointage.")
                with t_j:
                    jdf = get_justifications()
                    if len(jdf)>0: st.dataframe(jdf,use_container_width=True,height=260)
                    else: st.info("Aucune justification.")

st.markdown("---")
st.caption(f"Plateforme Formation · v10.0 · {'🟢 Supabase' if DB else '⚠️ Mémoire'} · "
           f"Vue : {'publique' if VUE_PUBLIC else 'direction'} · "
           f"Notes : {'visibles' if notes_visibles else 'masquées (vue publique)'}")
